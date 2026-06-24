#!/usr/bin/env python3
"""导出飞书多维表格初始数据（发版项 / 原声关联 / 触达清单）。

基于 release_perception_review.json + tickets/feedback 关键词召回。

用法:
  python3 scripts/export_feishu_voice_closure.py
  python3 scripts/export_feishu_voice_closure.py --perception auto   # 🟢=是 ⚪=否 🟡待确认
  python3 scripts/export_feishu_voice_closure.py --ledger ~/Downloads/发版项台账.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path

WORKSPACE = Path(__file__).resolve().parents[1]
REVIEW_PATH = WORKSPACE / 'data' / 'release_perception_review.json'
KEYWORDS_PATH = WORKSPACE / 'data' / 'feature_keywords.json'
TICKETS_PATH = WORKSPACE / 'data' / 'tickets.json'
FEEDBACK_PATH = WORKSPACE / 'data' / 'feedback.json'
OUT_DIR = WORKSPACE / 'exports' / 'feishu'
DATA_OUT = WORKSPACE / 'data'
LEDGER_JSON = DATA_OUT / 'voice_closure_ledger.json'
LINKS_JSON = DATA_OUT / 'voice_closure_links.json'
STATE_JSON = DATA_OUT / 'voice_closure_state.json'

# 发版项名称 → 额外搜索词（优先匹配 feature_keywords.json）
EXTRA_KW: dict[str, list[str]] = {}


def load_keywords() -> dict[str, list[str]]:
    if KEYWORDS_PATH.exists():
        return json.loads(KEYWORDS_PATH.read_text(encoding='utf-8'))
    return {}


def keywords_for_feature(
    name: str,
    kw_map: dict[str, list[str]],
    value_hint: str = '',
) -> list[str]:
    name = name.strip()
    kws: list[str] = []
    # 精确或包含匹配词典 key
    for key, words in kw_map.items():
        if key in name or name in key:
            kws.extend(words)
    # 从名称拆词
    kws.append(name)
    # 去人名后缀（石墨常见）
    short = re.sub(r'[周孟张王李陈荆韩胡汤操龙邢].{1,2}$', '', name).strip()
    if short and short != name:
        kws.append(short)
    # 引号内短语（如「名校专区」）
    for m in re.findall(r'[「『]([^」』]+)[」』]', name):
        kws.append(m)
    # 取中文片段 >=2
    for m in re.findall(r'[\u4e00-\u9fa5]{2,}', name + ' ' + (value_hint or '')):
        if len(m) >= 2:
            kws.append(m)
    # 去重保序
    seen: set[str] = set()
    out: list[str] = []
    for w in kws:
        w = w.strip()
        if not w or w in seen or len(w) < 2:
            continue
        seen.add(w)
        out.append(w)
    return out[:20]


def perception_label(auto: str, mode: str) -> str:
    if mode == 'auto':
        return {'suggested_yes': '是', 'suggested_no': '否', 'uncertain': '待确认'}.get(auto, '待确认')
    return '待确认'


def score_text(text: str, keywords: list[str]) -> int:
    text = (text or '').lower()
    score = 0
    hits = 0
    for kw in keywords:
        if kw.lower() in text:
            hits += 1
            score += max(2, min(len(kw), 8))
    # 多词共现加分，避免长词门槛过高
    if hits >= 2:
        score += hits
    return score


def item_id(version: str, name: str) -> str:
    return f'{version}::{name}'


def load_ledger_csv(path: Path) -> list[dict]:
    """读取飞书导出的发版项台账 CSV（保留用户确认的「用户可感知」等列）。"""
    with path.open(encoding='utf-8-sig', newline='') as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        row['关联原声数'] = 0
        if row.get('用户可感知') is None:
            row['用户可感知'] = ''
    return rows


def export_release_items(review: dict, kw_map: dict, perception_mode: str) -> list[dict]:
    rows: list[dict] = []
    for rel in review['releases']:
        for it in rel['items']:
            pid = item_id(rel['version'], it['name'])
            percep = perception_label(it['auto'], perception_mode)
            rows.append({
                '发版项ID': pid,
                '版本号': rel['version'],
                '发版日期': rel['date'],
                '机型': rel['model'],
                '发版项名称': it['name'],
                '来源': '石墨' if it['source'] == 'shimo' else 'Confluence',
                '自动初筛': {'suggested_yes': '建议可感知', 'suggested_no': '建议内部', 'uncertain': '待确认'}[it['auto']],
                '用户可感知': percep,
                '用户价值一句话': '',
                '关联原声数': 0,
                '备注': it.get('note') or '',
            })
    return rows


def write_workbook(ledger: list[dict], links: list[dict], outreach: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheets = [
        ('发版项台账', ledger),
        ('用户原声关联', links),
        ('触达清单', outreach),
    ]
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for r_idx, row in enumerate(rows, 2):
            for c_idx, h in enumerate(headers, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(h, ''))
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = 'A2'
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def collect_voices() -> list[dict]:
    voices: list[dict] = []
    if TICKETS_PATH.exists():
        for t in json.loads(TICKETS_PATH.read_text(encoding='utf-8')):
            text = ' '.join(filter(None, [t.get('problem'), t.get('desc'), t.get('classify')]))
            voices.append({
                '原声ID': f"ticket-{t.get('id')}",
                '来源': '工单',
                '用户ID': '',
                '设备': t.get('device') or '',
                '时间': t.get('time') or '',
                '原声摘要': (t.get('problem') or t.get('desc') or '')[:300],
                '链接': t.get('url') or '',
                '_text': text,
            })
    if FEEDBACK_PATH.exists():
        for r in json.loads(FEEDBACK_PATH.read_text(encoding='utf-8')):
            if len(r) < 6:
                continue
            voices.append({
                '原声ID': f"feedback-{r[0]}-{r[5]}",
                '来源': '用户反馈',
                '用户ID': str(r[0]),
                '设备': r[3] if len(r) > 3 else '',
                '时间': r[5],
                '原声摘要': (r[1] or '')[:300],
                '链接': f'https://cupid.zhenguanyu.com/#/analysis/user-feedback-list?userId={r[0]}',
                '_text': f"{r[1]} {r[2]} {r[3]} {r[4]}",
            })
    return voices


def export_voice_links(
    release_rows: list[dict],
    voices: list[dict],
    kw_map: dict,
    min_score: int = 2,
) -> list[dict]:
    perceivable = [r for r in release_rows if r['用户可感知'] == '是']
    links: list[dict] = []
    for feat in perceivable:
        kws = keywords_for_feature(
            feat['发版项名称'],
            kw_map,
            feat.get('用户价值一句话') or '',
        )
        if not kws:
            continue
        scored = []
        for v in voices:
            s = score_text(v['_text'], kws)
            if s >= min_score:
                scored.append((s, v))
        scored.sort(key=lambda x: (-x[0], x[1]['时间'] or ''))
        for s, v in scored:
            links.append({
                '关联ID': f"{feat['发版项ID']}::{v['原声ID']}",
                '发版项ID': feat['发版项ID'],
                '版本号': feat['版本号'],
                '发版项名称': feat['发版项名称'],
                '原声ID': v['原声ID'],
                '原声来源': v['来源'],
                '用户ID': v['用户ID'],
                '设备': v['设备'],
                '原声时间': v['时间'],
                '原声摘要': v['原声摘要'],
                '原声链接': v['链接'],
                '匹配分': s,
                '匹配关键词': '、'.join(k for k in kws if k.lower() in (v['_text'] or '').lower())[:80],
                '关联类型': '待确认',
                '关联已确认': '否',
                '需要触达': '待判断',
                '触达状态': '未触达',
                '触达话术': '',
                '触达人': '',
                '触达时间': '',
                '备注': '自动召回，请人工确认',
            })
        feat['关联原声数'] = sum(1 for l in links if l['发版项ID'] == feat['发版项ID'])
    return links


def export_outreach(links: list[dict]) -> list[dict]:
    """触达清单 = 关联表中「关联已确认=是」且「需要触达=是」的子集视图（初始导出候选）。"""
    rows = []
    for l in links:
        if l['原声来源'] != '用户反馈' or not l['用户ID']:
            continue
        rows.append({
            '关联ID': l['关联ID'],
            '版本号': l['版本号'],
            '发版项名称': l['发版项名称'],
            '用户ID': l['用户ID'],
            '设备': l['设备'],
            '原声摘要': l['原声摘要'],
            '原声链接': l['原声链接'],
            '触达状态': l['触达状态'],
            '触达话术': l['触达话术'],
            '触达人': l['触达人'],
            '触达时间': l['触达时间'],
            '备注': '候选（需先在关联表确认）',
        })
    return rows


def ledger_row_to_api(row: dict) -> dict:
    return {
        'id': row['发版项ID'],
        'version': row['版本号'],
        'date': row['发版日期'],
        'model': row['机型'],
        'name': row['发版项名称'],
        'source': row['来源'],
        'autoScreen': row.get('自动初筛') or '',
        'perceivable': row.get('用户可感知') == '是',
        'valueSummary': row.get('用户价值一句话') or '',
        'linkCount': int(row.get('关联原声数') or 0),
        'note': row.get('备注') or '',
    }


def link_row_to_api(row: dict) -> dict:
    return {
        'id': row['关联ID'],
        'featureId': row['发版项ID'],
        'version': row['版本号'],
        'featureName': row['发版项名称'],
        'voiceId': row['原声ID'],
        'voiceSource': row['原声来源'],
        'userId': row['用户ID'],
        'device': row['设备'],
        'voiceTime': row['原声时间'],
        'summary': row['原声摘要'],
        'url': row['原声链接'],
        'score': int(row['匹配分']),
        'keywords': row['匹配关键词'],
    }


def write_dashboard_json(
    ledger: list[dict],
    links: list[dict],
    min_score: int,
    source: str,
) -> None:
    DATA_OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ledger_doc = {
        'generated_at': now,
        'source': source,
        'items': [ledger_row_to_api(r) for r in ledger],
    }
    links_doc = {
        'generated_at': now,
        'min_score': min_score,
        'count': len(links),
        'links': [link_row_to_api(r) for r in links],
    }
    LEDGER_JSON.write_text(json.dumps(ledger_doc, ensure_ascii=False, indent=2), encoding='utf-8')
    LINKS_JSON.write_text(json.dumps(links_doc, ensure_ascii=False, indent=2), encoding='utf-8')
    if not STATE_JSON.exists():
        STATE_JSON.write_text(
            json.dumps({'updated_at': now, 'links': {}}, ensure_ascii=False, indent=2),
            encoding='utf-8',
        )
    print(f'看板 JSON: {LEDGER_JSON}')
    print(f'看板 JSON: {LINKS_JSON}')


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        path.write_text('', encoding='utf-8-sig')
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--perception', choices=['auto', 'all_pending'], default='auto')
    parser.add_argument('--ledger', type=Path, help='飞书导出的发版项台账 CSV（优先于自动初筛）')
    parser.add_argument('--min-score', type=int, default=2, help='匹配分门槛，默认 2；设为 0 可输出所有含任一关键词的命中')
    parser.add_argument('--json-out', action='store_true', help='同时输出 data/voice_closure_*.json 供发版看板使用')
    parser.add_argument('--json-only', action='store_true', help='仅输出看板 JSON，不写 CSV/Excel')
    args = parser.parse_args()

    kw_map = load_keywords()
    if args.ledger:
        release_rows = load_ledger_csv(args.ledger)
        log_src = f'飞书台账 {args.ledger}'
    else:
        review = json.loads(REVIEW_PATH.read_text(encoding='utf-8'))
        release_rows = export_release_items(review, kw_map, args.perception)
        log_src = 'release_perception_review.json'
    voices = collect_voices()
    link_rows = export_voice_links(release_rows, voices, kw_map, args.min_score)
    outreach_rows = export_outreach(link_rows)

    if args.json_out or args.json_only:
        write_dashboard_json(release_rows, link_rows, args.min_score, log_src)

    if not args.json_only:
        ts = datetime.now().strftime('%Y%m%d')
        write_csv(OUT_DIR / f'1_发版项台账_{ts}.csv', release_rows)
        write_csv(OUT_DIR / f'2_用户原声关联_{ts}.csv', link_rows)
        write_csv(OUT_DIR / f'3_触达清单候选_{ts}.csv', outreach_rows)
        xlsx_path = OUT_DIR / '学练机发版用户声音闭环.xlsx'
        write_workbook(release_rows, link_rows, outreach_rows, xlsx_path)

    yes_count = sum(1 for r in release_rows if r['用户可感知'] == '是')
    print(f'数据源: {log_src}')
    print(f'发版项: {len(release_rows)}（用户可感知 {yes_count}）')
    print(f'原声关联候选: {len(link_rows)}')
    print(f'触达候选(userId): {len(outreach_rows)}')
    if not args.json_only:
        print(f'输出目录: {OUT_DIR}')
        print(f'Excel: {xlsx_path}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
